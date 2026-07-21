# openpyxl is fine
# pyxlsb2 needed modifications from the official release

import pyxlsb2
import openpyxl
import os
import traceback

def cell_name(row, col):
    return "%s%d" % (openpyxl.utils.cell.get_column_letter(col), row)

def parse_xlsb(ffn, ofn, show_values):
    print("parse_xlsb(%s, %s)" % (ffn, ofn))
    print("writing output to %s" % ofn)
    
    with open(ofn, "w") as fp:
        with pyxlsb2.open_workbook(ffn) as wb:
            for sr in wb.sheets:
                sname = sr.name
                sheet = wb.get_sheet_by_name(sname)
                num_empty_rows = 0
                for row in sheet.rows():
                    has_content = False
                    for cell in row:
                        try:
                            if cell.value is None and cell.formula is None:
                                continue
                            has_content = True
                            if cell.formula is None or len(cell.formula) == 0:
                                if show_values:
                                    fp.write("%s!%s:{%s}\n" % (sname, cell_name(row.num+1, cell.col+1), cell.value))
                            else:
                                formula = pyxlsb2.formula.Formula.parse(cell.formula, cell.row, cell.col).stringify(wb)
                                if formula.startswith("RC("): continue
                                fp.write("%s!%s=%s\n" % (sname, cell_name(row.num+1, cell.col+1), formula))
                        except Exception as e:
                            fp.write("%s!%s: ERROR: %s\n" % (sname, cell_name(row.num+1, cell.col+1), str(e)))

                    if not has_content:
                        num_empty_rows += 1
                        if num_empty_rows >= 100: break
                    else:
                        num_empty_rows = 0

def parse_xls(ffn, ofn, show_values):
    print("parse_xls(%s, %s)" % (ffn, ofn))
    print("writing output to %s" % ofn)

    with open(ofn, "w") as fp:

        # openpyxl workbooks do not support context managers
        wb = None
        try:
            wb = openpyxl.load_workbook(ffn, read_only=True)
            for sname in wb.sheetnames:
                sheet = wb[sname]
                num_empty_rows = 0
                for row in sheet.rows:
                    has_content = False
                    for cell in row:
                        try:
                            # in openpyxl cell.value can be a formula object or a value
                            if cell.value is None: continue
                            has_content = True

                            # we cannot just cast to string since not all cell values have a string representation
                            if hasattr(cell.value, "text"):
                                value = cell.value.text
                            else:
                                value = str(cell.value)

                            if value.startswith("="):
                                formula = value[1:]
                                formula = formula.replace("_xll.", "")
                                if formula == "": continue
                                fp.write("%s!%s=%s\n" % (sname, cell_name(cell.row, cell.column), formula))
                            elif show_values:
                                fp.write("%s!%s:{%s}\n" % (sname, cell_name(cell.row, cell.column), value))
                        except Exception as e:
                            fp.write("%s!%s: ERROR: %s\n" % (sname, cell_name(cell.row, cell.column), str(e)))

                    if not has_content:
                        num_empty_rows += 1
                        if num_empty_rows >= 100: break
                    else:
                        num_empty_rows = 0
        finally:
            if wb is not None: wb.close()

def parse_fn(ffn, odn, show_values):
    try:
        fn = os.path.basename(ffn)
        ofn = os.path.join(odn, "%s.txt" % fn)
        if fn.startswith("~"):
            print("ignoring file starting with ~:", fn)
        elif fn.endswith(".xlsb"):
            parse_xlsb(ffn, ofn, show_values)
        elif fn.endswith(".xls") or fn.endswith(".xlsx") or fn.endswith(".xlsm"):
            parse_xls(ffn, ofn, show_values)
        else:
            print("ignoring file with unknown extension:", fn)
    except Exception as e:
        efn = os.path.join(odn, "error.log")
        with open(efn, "a") as efp:
            efp.write("Error processing %s: %s\n" % (ffn, str(e)))
            efp.write("%s\n" % traceback.format_exc())

def parse_dn(dn, odn, show_values):
    import glob

    fns = glob.glob(os.path.join(dn, '*.xls*'))
    for ffn in fns:
        parse_fn(ffn, odn, show_values)

def main(fns, odn=".", show_values=False):
    if len(fns) == 0:
        raise Exception("No files or directories provided")

    if not os.path.isdir(odn):
        print("creating directory %s" % odn)
        os.makedirs(odn)

    for fn in fns:
        if os.path.isdir(fn):
            dn = fn
            parse_dn(dn, odn, show_values)
        elif os.path.isfile(fn):
            parse_fn(fn, odn, show_values)

if __name__ == "__main__":
    import getopt
    import sys

    if len(sys.argv) == 1:
        print("running %s with no parameters" % sys.argv[0])
        print("current directory:", os.getcwd())
        print()
        while True:
            arg = input("enter argument (including options): ")
            if not arg: break
            sys.argv.append(arg)

    kwargs = {}
    opts, args = getopt.getopt(sys.argv[1:], "o:v")

    fns = args[:]
    if len(fns) == 0: fns = ["."]

    for opt in opts:
        if opt[0] == "-o": kwargs["odn"] = opt[1]
        elif opt[0] == "-v": kwargs["show_values"] = True

    main(fns, **kwargs)

